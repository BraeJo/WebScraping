import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24, italic=False, bold=True)
#Alternatively
fontobject = Font(name='Times New Roman', size=24, italic=False, bold=True)
ws['A1'].font = fontobject

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'
ws.merge_cells('A1:B1')
#unmerge cells
#ws. unmerge_cells('A1:B1')
ws['B2' ] = 450
ws['B3'] = 225
ws['B4'] = 150
ws['A8'] = 'Total'
ws['A8'].font = Font(size=16, bold=True)
ws['B8'] = 'SUM(B2:B4)'

ws.column_dimensions['A'].width = 25

wb.save('PythontoExcel.xlsx')